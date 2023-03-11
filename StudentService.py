import openpyxl
import mysql.connector
from Student import Student

listSV = []
mini = [7, 30]
gioi = 0
kha = 0
tb = 0


def getlistSV():
    return listSV


def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False


def SortCri(e):
    return e.MaSV


def tinhdtb(a, b, c):
    return round((a + b + c) / 3, 1)


def checkHK(e):
    if e >= 8:
        return "Giỏi"
    elif e >= 6.5:
        return "Khá"
    else:
        return "Trung Bình"


def max(i):
    max = 0
    if i == 0:
        for j in range(len(listSV)):
            if max < len(listSV[j].MaSV):
                max = len(listSV[j].MaSV)

    if i == 1:
        for j in range(len(listSV)):
            if max < len(listSV[j].Ho + " " + listSV[j].Ten):
                max = len(listSV[j].Ho + " " + listSV[j].Ten)

    if max < mini[i]:
        max = mini[i]

    return max


def convertdate(time):
    date = time.split("/")
    return date[2] + "-" + date[1] + "-" + date[0]


def revsconvertdate(time):
    date = time.split("-")
    return date[2] + "/" + date[1] + "/" + date[0]
def clearlist():
    return listSV.clear()


def readexcel():
    global hk
    global filename
    hk = [0, 0, 0]
    filename = input("Nhập tên file cần đọc:") + '.xlsx'
    book = openpyxl.load_workbook(filename)
    sh = book.active
    maxrow = sh.max_row
    i = 1
    t = "A" + str(i)
    while sh[t].value != "STT" or i == maxrow:
        i += 1
        t = "A" + str(i)
    if i == maxrow:
        print("Không đọc dược dữ liệu")
    else:
        global index
        index = i
        j = i + 1
        t = "A" + str(j)
        while type(sh[t].value) == int:

            stu = Student(sh.cell(row=j, column=2).value,
                          sh.cell(row=j, column=3).value,
                          sh.cell(row=j, column=4).value,
                          sh.cell(row=j, column=5).value,
                          round(float(sh.cell(row=j, column=6).value), 1),
                          round(float(sh.cell(row=j, column=7).value), 1),
                          round(float(sh.cell(row=j, column=8).value), 1),
                          0, ""
                          )
            stu.DTB = tinhdtb(stu.Toan, stu.Ly, stu.Hoa)
            stu.HK = checkHK(stu.DTB)
            if stu.HK == "Giỏi":
                hk[0] += 1
            elif stu.HK == "Khá":
                hk[1] += 1
            else:
                hk[2] += 1

            listSV.append(stu)
            j += 1
            t = "A" + str(j)
        listSV.sort(key=SortCri)


def readmysql():
    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
        cur.execute("SELECT * FROM student")
        global rows
        rows = cur.fetchall()

        for row in rows:
            stu = Student(row[1], row[2], row[3], revsconvertdate(str(row[4])), row[5], row[6], row[7], row[8], row[9])
            listSV.append(stu)
        listSV.sort(key=SortCri)
        print('Total Row(s):', cur.rowcount)
    except mysql.connector.Error as e:
        print(e)

    finally:
        cur.close()
        myconn.close()


def showsv(list):
    totallen = 0
    for i in range(2):
        totallen += max(i)

    print("_" * (totallen + 52))
    print(
        "| {} | {} | {:^10} | {:^5} | {:^5} | {:^5} | {:^5} |".format("STT".center(max(0)), "Họ và Tên".center(max(1)),
                                                                      "Ngay Sinh", "Toán", "Lý",
                                                                      "Hóa", "ĐTB", "HK"))
    for i in list:
        print(
            "| {:} | {:} | {:^10} | {:^5} | {:^5} | {:^5} | {:^5} |".format(i.MaSV.center(max(0)),
                                                                            (i.Ho + " " + i.Ten).center(max(1)),
                                                                            i.NgaySinh, i.Toan, i.Ly, i.Hoa,
                                                                            i.DTB, i.HK))
    print("|" + "_" * (totallen + 50) + "|")


def showsvinsql():
    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
        cur.execute("SELECT * FROM student")
        global rows
        rows = cur.fetchall()

        print('Total Row(s):', cur.rowcount)
    except mysql.connector.Error as e:
        print(e)

    finally:
        cur.close()
        myconn.close()

    totallen = 0
    for i in range(2):
        totallen += max(i)

    print("_" * (totallen + 52))
    print(
        "| {} | {} | {:^10} | {:^5} | {:^5} | {:^5} | {:^5} |".format("STT".center(max(0)), "Họ và Tên".center(max(1)),
                                                                      "Ngay Sinh", "Toán", "Lý",
                                                                      "Hóa", "ĐTB", "HK"))
    for i in rows:
        print(
            "| {:} | {:} | {:^10} | {:^5} | {:^5} | {:^5} | {:^5} |".format(i[1].center(max(0)),
                                                                         (i[2] + " " + i[3]).center(max(1)),
                                                                         revsconvertdate(str(i[4])), i[5], i[6], i[7], i[8]))
    print("|" + "_" * (totallen + 50) + "|")


def sqlconn():
    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
    except mysql.connector.errors.ProgrammingError:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="")
        cur = myconn.cursor()
        cur.execute("CREATE DATABASE pythondb")


def search(masv):
    sql = """SELECT id  FROM Student 
            WHERE MaSV = %s 
            """
    data = [masv]
    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
        cur.execute(sql, data)
        row = cur.fetchone()
        return row[0]
    except mysql.connector.Error as error:
        print(error)
    finally:
        cur.close()
        myconn.close()


def checkexist(e):
    sql = "SELECT masv, COUNT(*) FROM student WHERE masv = %s "
    val = (e.MaSV,)
    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
        cur.execute(sql, val)
        results = cur.fetchone()
        if results[1] == 0:
            return True
        else:
            return False
    except:
        print("Lỗi lấy dữ liệu")

    finally:
        cur.close()
        myconn.close()


def insert(stu):
    if checkexist(stu):
        sql = "INSERT INTO student(masv, ho, ten, ngaysinh, toan, ly, hoa, dtb, hk)" \
            + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        val = (stu.MaSV, stu.Ho, stu.Ten, convertdate(stu.NgaySinh), stu.Toan, stu.Ly, stu.Hoa, stu.DTB, stu.HK)
        try:
            myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
            cur = myconn.cursor()
            cur.execute(sql, val)
            myconn.commit()
        except:
            myconn.rollback()
            print(cur.rowcount, "record inserted")
        finally:
            cur.close()
            myconn.close()
        print("Mã Sinh Viên {} đã thêm".format(stu.MaSV))

        #print("Mã Sinh Viên {} đã tồn tại!".format(stu.MaSV))


def insertall(list):
    try:
        for i in list:
            insert(i)
        print("Da them du lieu vao MySql")
    except:
        print("Co loi da xay ra")



def update_student(list, id):
    stu = list[id - 1]
    # Câu lệnh update dữ liệu
    query = """ UPDATE student
                SET masv = %s, ho = %s, ten = %s,
                    ngaysinh = %s, toan = %s, 
                    ly = %s, hoa = %s,
                    dtb = %s, hk = %s
                WHERE id = %s """

    data = (stu.MaSV, stu.Ho, stu.Ten, convertdate(stu.NgaySinh), stu.Toan, stu.Ly, stu.Hoa, stu.DTB, stu.HK,
            id)

    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
        cur.execute(query, data)
        myconn.commit()

    except mysql.connector.Error as error:
        print(error)

    finally:
        # Đóng kết nối
        cur.close()
        myconn.close()


def changedata(list, id):
    masv = str(input("Nhập Mã Sinh Viên ( init:" + list[id - 1].MaSV + "):")).strip()

    ho = input("Nhập Họ ( init:" + list[id - 1].Ho + "):")
    while ho.isnumeric() is True:
        ho = input("Nhập lại Họ ( init:" + list[id - 1].Ho + ": ")
    ho = str(ho.strip())

    ten = input("Nhập Tên ( init:" + list[id - 1].Ten + "):")
    while ten.isnumeric() is True:
        ten = input("Nhập lại Tên ( init:" + list[id - 1].Ten + ": ")
    ten = str(ten.strip())

    ngaysinh = input("Nhập Ngay Sinh ( init:" + list[id - 1].NgaySinh + ")(format: DD-MM-YYYY): ")

    toan = input("Nhập điểm Toán ( init:" + str(list[id - 1].Toan) + "):")
    while isfloat(toan) is False and toan.isnumeric() is False:
        toan = input("Nhập lại điểm Toán ( init:" + str(list[id - 1].Toan) + "): ")
    toan = round(float(toan), 1)

    ly = input("Nhập điểm Lý ( init:" + str(list[id - 1].Ly) + "):")
    while isfloat(ly) is False and ly.isnumeric() is False:
        ly = input("Nhập lại điểm Lý ( init:" + str(list[id - 1].Ly) + "): ")
    ly = round(float(ly), 1)

    hoa = input("Nhập điểm Hóa ( init:" + str(list[id - 1].Hoa) + "):")
    while isfloat(hoa) is False and hoa.isnumeric() is False:
        hoa = input("Nhập lại điểm Hóa ( init:" + str(list[id - 1].Hoa) + "): ")
    hoa = round(float(hoa), 1)

    dtb = tinhdtb(toan, ly, hoa)
    hk = checkHK(dtb)
    stu = Student(masv, ho, ten, ngaysinh, toan, ly, hoa, round(dtb, 2), hk)
    listSV[id - 1] = stu
    listSV.sort(key=SortCri)


def deletedata(id):
    query = "DELETE FROM student WHERE id = %s"

    try:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
        cur = myconn.cursor()
        cur.execute(query, (id,))
        myconn.commit()
        print("Đã xóa thành công!")

    except mysql.connector.Error as error:
        print(error)

    finally:
        # Đóng kết nối
        cur.close()
        myconn.close()
